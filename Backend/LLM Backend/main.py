from fastapi import FastAPI, HTTPException, File, Request
from pydantic import BaseModel
import httpx
import os
from dotenv import load_dotenv
from groq import Groq
from fastapi.responses import JSONResponse
import base64
import textract
import tempfile
from io import BytesIO
from openpyxl import load_workbook
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
import io
import json
from typing import Literal, List
import openai
from cerebras.cloud.sdk import Cerebras
from mem0 import AsyncMemoryClient
import contextvars

load_dotenv()

app = FastAPI()

GROQ_API_KEY = os.getenv("GROQ_API_KEY_1", "")
MISTRAL_API_KEY = os.getenv("MISTRAL_API_KEY", "")
HUGGINGFACE_API_KEY = os.getenv("HUGGINGFACE_API_KEY", "")
CLOUDFLARE_API_KEY = os.getenv("CLOUDFLARE_API_KEY", "")
CEREBRAS_API_KEY = os.getenv("CEREBRAS_API_KEY_1", "")
MEM0_API_KEY = os.getenv("MEM0_API_KEY", "")

client = Cerebras(
    api_key=os.environ.get("CEREBRAS_API_KEY_1"),
)


mem_client = AsyncMemoryClient(
    api_key=os.environ.get("MEM0_API_KEY"),
)


groqClient = Groq(
    api_key=GROQ_API_KEY,
)
groqClient_1 = Groq(
    api_key=os.environ.get("GROQ_API_KEY_2"),
)
providers = {
    "cerebras": {
        "base_url": "https://api.cerebras.ai/v1",
        "api_key": os.environ.get("CEREBRAS_API_KEY_1"),
        "api_key_2": os.environ.get("CEREBRAS_API_KEY_2"),
        "api_key_3": os.environ.get("CEREBRAS_API_KEY_3"),
    },
    "groq": {
        "base_url": "https://api.groq.com/openai/v1",
        "api_key_1": os.environ.get("GROQ_API_KEY_1"),
        "api_key_2": os.environ.get("GROQ_API_KEY_2"),
    },
    "openrouter": {
        "base_url": "https://openrouter.ai/api/v1",
        "api_key1": os.environ.get("OPENROUTER_API_KEY_1"),
        "api_key2": os.environ.get("OPENROUTER_API_KEY_2"),
        "api_key3": os.environ.get("OPENROUTER_API_KEY_3"),
    },
}


class ChatMessage(BaseModel):
    role: Literal["system", "user", "assistant", "function"]
    content: str


class LLMRequest(BaseModel):
    prompt: List[ChatMessage]
    model: str
    provider: list
    premium: bool
    is_agent: bool = False


class OcrRequest(BaseModel):
    imageBase64: str


class FileRequest(BaseModel):
    buffer: str
    name: str
    mime: str


current_user_id = contextvars.ContextVar("current_user_id", default=None)


@app.middleware("http")
async def set_user_context(request: Request, call_next):
    uid = request.headers.get("X-User-Id")
    token = current_user_id.set(uid)
    try:
        return await call_next(request)
    finally:
        current_user_id.reset(token)


@app.post("/llm")
async def llm_query(request: LLMRequest):
    prodiver = request.provider
    if "mistral" in prodiver:
        return await query_mistral(request)
    elif "cloudflare" in prodiver:
        return await query_cloudflare(request)
    else:
        return await providerRouting(request)


@app.post("/ocr")
async def ocr_query(req: OcrRequest):
    imageBase64 = req.imageBase64
    chat_completion = groqClient.chat.completions.create(
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": "Что на этом изображении? Опиши подробно"},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{imageBase64}",
                        },
                    },
                ],
            }
        ],
        model="meta-llama/llama-4-scout-17b-16e-instruct",
    )
    return chat_completion.choices[0].message.content


@app.post("/files")
async def files_recognize(req: FileRequest):
    data = base64.b64decode(req.buffer)
    mediaType = ""
    ext = req.name.rsplit(".", 1)[-1].lower()

    if ext == "pdf":
        rsrcmgr = PDFResourceManager()
        laparams = LAParams()
        retstr = io.StringIO()
        device = TextConverter(rsrcmgr, retstr, laparams=laparams)
        interpreter = PDFPageInterpreter(rsrcmgr, device)

        fp = BytesIO(data)
        fp.seek(0)

        total_text = ""
        mediaType = "pdf"
        for page in PDFPage.get_pages(fp, check_extractable=False):
            interpreter.process_page(page)
            page_text = retstr.getvalue()
            retstr.truncate(0)
            retstr.seek(0)

            remaining = 3000 - len(total_text)
            total_text += page_text[:remaining]
            if len(total_text) >= 3000:
                break

            device.close()
            retstr.close()
            fp.close()
            return total_text
    elif ext in ("xls", "xlsx"):
        wb = load_workbook(filename=BytesIO(data), read_only=True)
        parts = []
        for ws in wb.worksheets:
            parts.append(f"=== Sheet: {ws.title} ===")
            for row in ws.iter_rows(values_only=True):
                line = "\t".join("" if c is None else str(c) for c in row)
                parts.append(line)
        text = "\n".join(parts)
        mediaType = "table"
    elif ext in ("mp3", "wav", "ogg", "m4a"):
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp:
                tmp.write(data)
                tmp_path = tmp.name

            with open(tmp_path, "rb") as f:
                result = groqClient.audio.transcriptions.create(
                    model="whisper-large-v3",
                    file=("audio." + ext, f.read(), f"audio/{ext}"),
                    response_format="text",
                    prompt="Дословно напиши, точную расшифровку текста на том языке на котором аудио записано, и ничего не более, если ты не нашёл слов в аудио файле, то опиши  что за звуки там.",
                )
            mediaType = "audio"
            text = json.dumps(result, indent=2, ensure_ascii=False)
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.unlink(tmp_path)

    else:
        ext = req.name.rsplit(".", 1)[-1].lower()
        suffix = f".{ext}" if ext else ""
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp.write(data)
                tmp_path = tmp.name
            raw = textract.process(tmp_path)
            text = raw.decode("utf-8", errors="ignore")

        except Exception as e:
            return {"error": "Processing failed", "detail": str(e)}

        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
    return {"content": text, "type": mediaType}


async def query_mistral(request: LLMRequest):
    url = "https://api.mistral.ai/v1/chat/completions"
    headers = {"Authorization": f"Bearer {MISTRAL_API_KEY}"}
    payload = {
        "model": request.model,
        "messages": [m.dict() for m in request.prompt],
    }
    timeout = httpx.Timeout(
    connect=10.0,  
    read=120.0,   
    write=60.0,
    pool=10.0
)
    async with httpx.AsyncClient() as client:
        resp = await client.post(url, json=payload, headers=headers, timeout=timeout)
        if resp.status_code != 200:
            raise HTTPException(status_code=resp.status_code, detail=resp.text)
        data = resp.json()
        return JSONResponse(
            content={
                "type": "text",
                "content": data["choices"][0]["message"]["content"],
            }
        )


async def query_cloudflare(request: LLMRequest):
    url = (
        f"https://api.cloudflare.com/client/v4/accounts/"
        f"5054130e5a7ddf582ed30bdae4809a82/ai/run/@{request.model}"
    )
    headers = {"Authorization": f"Bearer {CLOUDFLARE_API_KEY}"}
    payload = {"prompt": request.prompt}

    timeout = httpx.Timeout(
        connect=60.0,
        read=120.0,
        write=60.0,
        pool=5.0,
    )
    async with httpx.AsyncClient(timeout=timeout) as client:
        try:
            resp = await client.post(url, json=payload, headers=headers)
        except httpx.ReadTimeout:
            raise HTTPException(
                504,
                detail="Upstream Read Timeout: модель не успела ответить за отведённое время",
            )
        except httpx.RequestError as e:
            raise HTTPException(502, detail=f"Upstream Request Error: {e}")

        if resp.status_code != 200:
            raise HTTPException(status_code=resp.status_code, detail=resp.text)

        raw_bytes = await resp.aread()

        encoded = base64.b64encode(raw_bytes).decode("ascii")
        return JSONResponse(content={"type": "image", "content": encoded})


import re
import datetime

from typing import Any, Dict, List, Optional

def add_prefix_if_first_is_system(full_messages: list, prefix: str) -> list:
    if not full_messages:
        return full_messages

    first = full_messages[0]

    if isinstance(first, dict):
        role = first.get("role")
        content = first.get("content", "") or ""
    else:
        role = getattr(first, "role", None)
        content = getattr(first, "content", "") or ""

    if not role or str(role).lower() != "system":
        return full_messages

    if content.startswith(prefix):
        return full_messages

    new_content = f"{prefix}{content}"

    if isinstance(first, dict):
        first["content"] = new_content
    else:
        try:
            setattr(first, "content", new_content)
        except Exception:
            full_messages[0] = {"role": role, "content": new_content}

    return full_messages


def extract_api_keys_from_provider_conf(conf: Dict[str, Any]) -> List[str]:
    keys = []
    for k, v in conf.items():
        kn = k.lower().replace("-", "_")
        if kn.startswith("api_key") or kn.startswith("api") or "key" in kn:
            if isinstance(v, str) and v.strip():
                keys.append(v.strip())
    return keys


def detect_status_from_exception(e: Exception) -> Optional[int]:
    for attr in ("http_status", "status_code", "code"):
        val = getattr(e, attr, None)
        if isinstance(val, int):
            return val
        if isinstance(val, str) and val.isdigit():
            return int(val)
    resp = getattr(e, "response", None)
    if resp is not None:
        sc = getattr(resp, "status_code", None) or getattr(resp, "status", None)
        try:
            if isinstance(sc, int):
                return sc
            if isinstance(sc, str) and sc.isdigit():
                return int(sc)
        except Exception:
            pass
    m = re.search(r"\b(4\d{2}|5\d{2})\b", str(e))
    if m:
        try:
            return int(m.group(1))
        except Exception:
            pass
    return None


def extract_headers_from_exception(e: Exception) -> Dict[str, str]:
    hdrs: Dict[str, str] = {}
    resp = getattr(e, "response", None)
    if resp is not None:
        try:
            raw = getattr(resp, "headers", None)
            if raw:
                for k, v in raw.items():
                    hdrs[k.lower()] = v
                return hdrs
        except Exception:
            pass
    try:
        raw = getattr(e, "headers", None)
        if raw:
            for k, v in raw.items():
                hdrs[k.lower()] = v
            return hdrs
    except Exception:
        pass
    return hdrs


def parse_retry_seconds_from_headers(hdrs: Dict[str, str]) -> Optional[float]:
    if not hdrs:
        return None

    def parse_time_value(v: str) -> Optional[float]:
        if not v:
            return None
        s = v.strip().lower()
        if re.match(r"^\d+(\.\d+)?$", s):
            return float(s)
        total = 0.0
        found = False
        for val, unit in re.findall(r"(\d+(?:\.\d+)?)([smh])", s):
            found = True
            fv = float(val)
            if unit == "s":
                total += fv
            elif unit == "m":
                total += fv * 60
            elif unit == "h":
                total += fv * 3600
        if found:
            return total
        try:
            from email.utils import parsedate_to_datetime
            dt = parsedate_to_datetime(v)
            if dt is not None:
                now = datetime.datetime.utcnow().replace(tzinfo=datetime.timezone.utc)
                delta = (dt - now).total_seconds()
                return max(delta, 0.0)
        except Exception:
            pass
        return None

    ra = hdrs.get("retry-after")
    if ra:
        val = parse_time_value(ra)
        if val is not None:
            return val

    for key in (
        "x-ratelimit-reset-tokens",
        "x-ratelimit-reset-requests",
        "x-ratelimit-reset-tokens-minute",
        "x-ratelimit-reset-requests-minute",
        "x-ratelimit-reset",
    ):
        v = hdrs.get(key)
        if v:
            val = parse_time_value(v)
            if val is not None:
                return val

    return None

import logging
from typing import Any, Dict, List, Optional
import asyncio

async def _call_completion_with_flex(client, model: str, messages: List[Dict[str, str]]):
    res = client.chat.completions.create(model=model, messages=messages)
    if asyncio.iscoroutine(res):
        return await res
    return await asyncio.to_thread(lambda: res)


async def providerRouting(request: LLMRequest):
    provider = providers[request.provider[0]]
    agent_use = 0

    if request.provider[0] != "groq" and request.provider[0] != 'openrouter':
        try:
            model = request.model.split("/", 1)[1]
        except Exception:
            model = request.model
    else:
        model = request.model

    think_remove_re = re.compile(r"(?is)<think\b[^>]*>.*?</think\s*>")

    modified_prompt = []
    for item in request.prompt:
        if isinstance(item, dict):
            role = item.get("role")
            content = item.get("content", "")
        else:
            role = getattr(item, "role", None)
            content = getattr(item, "content", "")

        if content is None:
            content = ""
        if not isinstance(content, str):
            content = str(content)

        new_content = think_remove_re.sub("", content).strip()
        new_content = re.sub(r"\n{3,}", "\n\n", new_content)

        if not role:
            role = "user"

        modified_prompt.append({"role": str(role), "content": new_content})

    if request.is_agent:
        full_messages = await tool_agent_call([m.copy() for m in modified_prompt])
        if not isinstance(full_messages, list):
            raise RuntimeError("tool_agent_call вернул не список сообщений")
        full_messages = [
            (
                {
                    "role": str(m.get("role", "user")),
                    "content": str(m.get("content", "")),
                }
                if isinstance(m, dict)
                else {"role": "user", "content": str(m)}
            )
            for m in full_messages
        ]
        agent_use += 1
    else:
        full_messages = modified_prompt

    prefix = (
        f"Сейчас {datetime.date.today().isoformat()}, это 100 процентно правильная дата, "
        "ориентируйся на неё. Если видишь, что ответ пришёл от tool или function, то доверяй ему на 99 процентов. "
        "Если в сообщениях роли function/tool содержится результат внешнего поиска, считай эти данные актуальными..."
    )
    full_messages = add_prefix_if_first_is_system(full_messages, prefix)

    try:
        safe_messages = sanitize_for_provider(full_messages)
    except Exception as e:
        print("Sanitization error:", repr(e))
        return JSONResponse(
            status_code=500,
            content={"error": "sanitization_error", "detail": str(e)},
        )

    api_keys = extract_api_keys_from_provider_conf(provider)
    if not api_keys:
        return JSONResponse(status_code=500, content={"error": "no_api_keys", "detail": "no api keys found for provider"})

    max_rounds = 20  
    per_key_backoff = 0.2  
    round_index = 0
    last_exception = None

    while round_index < max_rounds:
        round_index += 1
        saw_retry_header = False
        retry_seconds = None

        for idx, api_key in enumerate(api_keys):
            logging.debug("Trying provider %s key %d/%d (round %d)", request.provider[0], idx + 1, len(api_keys), round_index)
            openai_client = openai.OpenAI(base_url=provider["base_url"], api_key=api_key)
            try:
                completion = await _call_completion_with_flex(openai_client, model, safe_messages)

                resp_text = (
                    completion.choices[0].message.content
                    if getattr(completion, "choices", None)
                    else ""
                )
                return JSONResponse(
                    content={"type": "text", "content": resp_text},
                    headers={"Agent-Use": f"{agent_use}"},
                )

            except Exception as e:
                logging.exception("Provider call failed with key index %d: %s", idx, repr(e))
                last_exception = e
                status = detect_status_from_exception(e)
                headers = extract_headers_from_exception(e)

                if status in (429, 402):
                    logging.info("Detected status %s for key %d; rotating to next key", status, idx + 1)
                    await asyncio.sleep(per_key_backoff)
                    continue

                parsed = parse_retry_seconds_from_headers(headers)
                if parsed is not None:
                    saw_retry_header = True
                    retry_seconds = parsed
                    logging.info("Detected retry header, will wait %.2fs before next round", retry_seconds)
                    break

                if status is not None and status >= 500:
                    logging.info("Transient server error %s, try next key after short sleep", status)
                    await asyncio.sleep(per_key_backoff)
                    continue

                logging.error("Unrecoverable provider error: %s", repr(e))
                return JSONResponse(status_code=500, content={"error": "provider_error", "detail": str(e)})

        if saw_retry_header and retry_seconds is not None:
            wait_time = float(retry_seconds) + 0.5
            max_wait = 300.0
            if wait_time > max_wait:
                wait_time = max_wait
            logging.info("Waiting %.2fs due to rate-limit headers before retrying key pool", wait_time)
            await asyncio.sleep(wait_time)
            continue
        else:
            sleep_time = min(2 ** min(round_index, 6), 60)
            logging.info("All keys exhausted (no retry header). Sleeping %.2fs before next round", sleep_time)
            await asyncio.sleep(sleep_time)
            continue

    logging.error("providerRouting: max_rounds reached, failing")
    return JSONResponse(status_code=502, content={"error": "provider_unavailable", "detail": str(last_exception)})


from typing import Dict


def sanitize_for_provider(messages: List[Dict]) -> List[Dict]:

    out = []
    for i, m in enumerate(messages):
        role = (m.get("role") or "").lower()
        content = m.get("content", "") or ""
        base = {"role": role, "content": content}

        if role == "tool":
            tcid = m.get("tool_call_id") or m.get("toolCallId") or m.get("call_id")
            if tcid:
                nm = {"role": "tool", "content": content, "tool_call_id": str(tcid)}
                if m.get("name"):
                    nm["name"] = m.get("name")
                out.append(nm)
            else:

                assistant_content = content

                out.append({"role": "assistant", "content": assistant_content})
        else:
            out.append({"role": role, "content": content})
    return out


async def add_memory(memory: str, user_id: str | None = None):
    uid = user_id or current_user_id.get()
    if not uid:
        raise HTTPException(status_code=401, detail="user_id is missing")
    messages = [
        {"role": "user", "content": str(memory)},
    ]
    result = await mem_client.add(messages, user_id=uid, version="v2")
    return result


async def search_memory(query: str, user_id: str | None = None):
    uid = user_id or current_user_id.get()
    if not uid:
        raise HTTPException(status_code=401, detail="user_id is missing")

    filters = {"AND": [{"user_id": str(uid)}]}
    results = await mem_client.search(str(query), version="v2", filters=filters)

    try:
        sorted_results = sorted(
            results if isinstance(results, list) else [],
            key=lambda r: r.get("score", 0),
            reverse=True,
        )
        top3 = sorted_results[:3]
    except Exception:
        top3 = []

    return {"matches": top3}


tools = [
    {
        "type": "function",
        "function": {
            "name": "add_memory",
            "description": "Позволяет добавить воспоминание о пользователе. Используй тогда тогда когда об этом попросит пользователь или очень потребуется, например очень важная деталь о пользователе",
            "parameters": {
                "type": "object",
                "properties": {
                    "memory": {
                        "type": "string",
                        "description": "Строка с текстом которое сохраниться в воспоминание. ",
                    }
                },
                "required": ["memory"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_memory",
            "description": "Позволяет совершить поиск по воспоминаниям о пользователе. Используй тогда тогда когда об этом попросит пользователь или очень потребуется, например вспомнить важную информацию о пользователе, которой у тебя нету в контексте сообщений",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Строка с текстом по которому будет проведён поиск. Пиши сюда например, когда пользователь просит что-то вспомнить, но информации об этом ты не можешь найти в контексте диалога и прошлых сообщениях ",
                    }
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "ocr_tool",
            "description": "Распознаёт изображение Base64 и возвращает текст",
            "parameters": {
                "type": "object",
                "properties": {"image_b64": {"type": "string"}},
                "required": ["image_b64"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "files_tool",
            "description": "Извлекает текст из файла (pdf, xlsx, прочие)",
            "parameters": {
                "type": "object",
                "properties": {
                    "buffer": {"type": "string"},
                    "name": {"type": "string"},
                },
                "required": ["buffer", "name"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "web_search",
            "description": "Даёт доступ к поиску в интернетею. Доверяй этому инструменту на 95 процентов, варианты тут не ошибочны и основанны на новостях и реальной ситуации",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Строка с поисковым запросом",
                    },
                },
                "required": ["query"],
            },
        },
    },
        {
        "type": "function",
        "function": {
            "name": "python_code_execution",
            "description": "Позволяет тебе исполнить почти любой python code, используй когда нужно что-то проверить, посчитать очень большие числа, где важна точность, ТОЛЬКО ОТ СЧЁТА НА МИЛЛЛИАРДЫ или посмотреть выполняется ли python код ли  правильно",
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": "Строка с кодом для выполнения",
                    },
                },
                "required": ["code"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "science_search",
            "description": r"""
                Используй этот инструмент тогда когда тебе нужны тяжелые вычисления в математики, физике или  химии, или ответы на очень тяжелые вопросы из тех же сфер, вот памятка по работе.- WolframAlpha understands natural language queries about entities in chemistry, physics, geography, history, art, astronomy, and more.
                - WolframAlpha performs mathematical calculations, date and unit conversions, formula solving, etc.`
                - Convert inputs to simplified keyword queries whenever possible (e.g. convert "how many people live in France" to "France population").
                - Send queries in English only; translate non-English queries before sending, then respond in the original language.
                - Display image URLs with Markdown syntax: ![URL]
                - ALWAYS use this exponent notation: `6*10^14`, NEVER `6e14`.
                - ALWAYS use {"input": query} structure for queries to Wolfram endpoints; `query` must ONLY be a single-line string.
                - ALWAYS use proper Markdown formatting for all math, scientific, and chemical formulas, symbols, etc.:  '$$\n[expression]\n$$' for standalone cases and '\( [expression] \)' when inline.
                - Never mention your knowledge cutoff date; Wolfram may return more recent data.
                - Use ONLY single-letter variable names, with or without integer subscript (e.g., n, n1, n_1).
                - Use named physical constants (e.g., 'speed of light') without numerical substitution.
                - Include a space between compound units (e.g., "Ω m" for "ohm*meter").
                - To solve for a variable in an equation with units, consider solving a corresponding equation without units; exclude counting units (e.g., books), include genuine units (e.g., kg).
                - If data for multiple properties is needed, make separate calls for each property.
                - If a WolframAlpha result is not relevant to the query:
                -- If Wolfram provides multiple 'Assumptions' for a query, choose the more relevant one(s) without explaining the initial result. If you are unsure, ask the user to choose.
                -- Re-send the exact same 'input' with NO modifications, and add the 'assumption' parameter, formatted as a list, with the relevant values.
                -- ONLY simplify or rephrase the initial query if a more relevant 'Assumption' or other input suggestions are not provided.
                -- Do not explain each step unless user input is needed. Proceed directly to making a better API call based on the available assumptions.
                """,
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Строка с запросом для api",
                    }
                },
                "required": ["query"],
            },
        },
    },
]


async def web_search(query: str):
    response = groqClient.chat.completions.create(
        model="compound-beta",
        messages=[
            {
                "role": "user",
                "content": f"Проведи поиск по интернету и найди ответ на вопрос: {query}",
            }
        ],
    )
    message = response.choices[0].message

    executed_tools = getattr(message, "executed_tools", None)
    if not executed_tools:
        return [message.content.strip()]

    tool0 = executed_tools[0]
    search_results = getattr(tool0, "search_results", None)
    if not search_results:
        return [message.content.strip()]

    results = None

    if hasattr(search_results, "results"):
        results = getattr(search_results, "results")
    else:
        try:
            sr_dict = search_results.dict()
        except Exception:
            sr_dict = None

        if isinstance(sr_dict, dict):
            results = sr_dict.get("results", [])
        else:
            try:
                results = dict(search_results).get("results", [])
            except Exception:
                results = []

    results = results or []
    top3 = results[:3]

    top3_text = ""
    for r in top3:
        if isinstance(r, dict):
            title = r.get("title", "—")
            url = r.get("url", "—")
            snippet = (r.get("content", "") or "")[:200]
        else:
            title = getattr(r, "title", "—")
            url = getattr(r, "url", "—")
            snippet = (getattr(r, "content", "") or "")[:200]

        top3_text = f"{top3_text}\n - {title}\n -{url}  {snippet}… "

    if not top3_text:
        return [message.content.strip()]
    return top3_text


async def python_code_execution(code: str):
    response = groqClient.chat.completions.create(
        model="compound-beta",
        messages=[
            {
                "role": "user",
                "content": f"Исполни данный далее тебе Python код и напиши, что он вывел: {code}",
            }
        ],
    )
    return response.choices[0].message.content

async def science_search(query: str):
    url = f"https://api.wolframalpha.com/v2/query?appid=TV3TVAVWAR&input={query}&output=JSON"
    timeout = httpx.Timeout(connect=60.0, read=120.0, write=60.0, pool=5.0)

    async with httpx.AsyncClient(timeout=timeout) as client:
        resp = await client.get(url)
        text = resp.text
        if resp.status_code != 200:
            raise HTTPException(status_code=resp.status_code, detail=text)
        try:
            return resp.json()
        except json.JSONDecodeError:
            raise HTTPException(
                status_code=502, detail=f"Invalid JSON from Wolfram: {text[:200]}..."
            )


available_functions = {
    "ocr_tool": ocr_query,
    "files_tool": files_recognize,
    "web_search": web_search,
    "science_search": science_search,
    "add_memory": add_memory,
    "search_memory": search_memory,
    "python_code_execution": python_code_execution,
}


async def tool_agent_call(start_messages: ChatMessage):
    messages: List[dict] = start_messages

    while True:
        resp = client.chat.completions.create(
            model="qwen-3-32b",
            messages=messages,
            tools=tools,
            tool_choice="auto",
            temperature=0.1,
        )
        msg = resp.choices[0].message
        if not msg.tool_calls:
            break

        messages.append(msg.model_dump())

        call = msg.tool_calls[0]
        fname = call.function.name

        if fname not in available_functions:
            raise ValueError(f"Unknown tool requested: {fname!r}")

        args_dict = json.loads(call.function.arguments)
        output = await available_functions[fname](**args_dict)

        messages.append(
            {
                "role": "tool",
                "tool_call_id": call.id,
                "name": fname,
                "content": json.dumps(output),
            }
        )
    return messages
